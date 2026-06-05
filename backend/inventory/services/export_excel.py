from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def style_header(row):
    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = max_length + 3


def format_currency_cell(cell):
    cell.number_format = '#,##0"đ"'


def create_sales_report_workbook(invoices):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao ban hang"

    ws["A1"] = "BÁO CÁO BÁN HÀNG"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:J1")

    headers = [
        "STT",
        "Mã hóa đơn",
        "Ngày bán",
        "Khách hàng",
        "Mã hàng",
        "Tên sản phẩm",
        "Size",
        "Màu",
        "Số lượng",
        "Thành tiền",
    ]

    ws.append([])
    ws.append(headers)
    style_header(ws[3])

    row_index = 1

    for invoice in invoices:
        for item in invoice.items.all():
            ws.append([
                row_index,
                invoice.invoice_code,
                invoice.sale_date.strftime("%d/%m/%Y"),
                invoice.customer.name if invoice.customer else "Khách lẻ",
                item.product_variant.product.code,
                item.product_variant.product.name,
                item.product_variant.size,
                item.product_variant.color,
                item.quantity,
                float(item.subtotal),
            ])

            format_currency_cell(ws.cell(row=ws.max_row, column=10))
            row_index += 1

    auto_fit_columns(ws)
    return wb


def create_import_report_workbook(receipts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao nhap hang"

    ws["A1"] = "BÁO CÁO NHẬP HÀNG"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:J1")

    headers = [
        "STT",
        "Mã phiếu nhập",
        "Ngày nhập",
        "Nhà cung cấp",
        "Mã hàng",
        "Tên sản phẩm",
        "Size",
        "Màu",
        "Số lượng",
        "Thành tiền",
    ]

    ws.append([])
    ws.append(headers)
    style_header(ws[3])

    row_index = 1

    for receipt in receipts:
        for item in receipt.items.all():
            ws.append([
                row_index,
                receipt.receipt_code,
                receipt.import_date.strftime("%d/%m/%Y"),
                receipt.supplier.name,
                item.product_variant.product.code,
                item.product_variant.product.name,
                item.product_variant.size,
                item.product_variant.color,
                item.quantity,
                float(item.subtotal),
            ])

            format_currency_cell(ws.cell(row=ws.max_row, column=10))
            row_index += 1

    auto_fit_columns(ws)
    return wb


def create_inventory_report_workbook(variants):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao ton kho"

    ws["A1"] = "BÁO CÁO TỒN KHO"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:K1")

    headers = [
        "STT",
        "Danh mục",
        "Mã hàng",
        "Tên sản phẩm",
        "Size",
        "Màu",
        "Barcode",
        "Giá nhập",
        "Giá bán",
        "Tồn kho",
        "Ngưỡng cảnh báo",
    ]

    ws.append([])
    ws.append(headers)
    style_header(ws[3])

    for index, variant in enumerate(variants, start=1):
        ws.append([
            index,
            variant.product.category.name,
            variant.product.code,
            variant.product.name,
            variant.size,
            variant.color,
            variant.barcode,
            float(variant.import_price),
            float(variant.sale_price),
            variant.current_stock,
            variant.low_stock_threshold,
        ])

        format_currency_cell(ws.cell(row=ws.max_row, column=8))
        format_currency_cell(ws.cell(row=ws.max_row, column=9))

    auto_fit_columns(ws)
    return wb